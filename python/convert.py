"""
PDF conversion sidecar for the PDF Editor's Rust server.

CLI contract:
    python convert.py --mode docx|xlsx --input <input.pdf> --output <output-file> [--pages 0,2,5]

On success: writes the output file, prints a single JSON line to stdout:
    {"ok": true, "pages": <number of pages converted>}
and exits 0.

On failure: prints a single JSON line to stderr:
    {"ok": false, "error": "<human-readable message>"}
and exits 1. The JSON line is always the LAST line written to stderr.
"""

import argparse
import io
import json
import logging
import os
import sys


class ConvertError(Exception):
    """Raised for any validation/conversion failure with a human-readable message."""


def parse_pages_arg(pages_arg, page_count):
    """Parse the --pages string into a sorted list of validated, unique 0-based indices.

    Omitted (None) means "all pages".
    """
    if pages_arg is None:
        return list(range(page_count))

    raw_parts = [p.strip() for p in pages_arg.split(",")]
    raw_parts = [p for p in raw_parts if p != ""]
    if not raw_parts:
        raise ConvertError("--pages given but empty")

    indices = []
    seen = set()
    for part in raw_parts:
        try:
            idx = int(part)
        except ValueError:
            raise ConvertError(f"invalid page index '{part}' (must be an integer)")
        if idx in seen:
            raise ConvertError(f"duplicate page index {idx}")
        seen.add(idx)
        if idx < 0 or idx >= page_count:
            raise ConvertError(
                f"page index {idx} out of range (document has {page_count} pages)"
            )
        indices.append(idx)

    return indices


def open_pdf_for_validation(input_path):
    """Open the PDF with PyMuPDF to get page count and check encryption.

    Returns the page count. Raises ConvertError on failure or if encrypted.
    """
    import fitz  # PyMuPDF

    try:
        doc = fitz.open(input_path)
    except Exception as exc:
        raise ConvertError(f"failed to open PDF: {exc}")

    try:
        if doc.needs_pass or doc.is_encrypted:
            raise ConvertError("document is encrypted")
        page_count = doc.page_count
    finally:
        doc.close()

    if page_count == 0:
        raise ConvertError("document has no pages")

    return page_count


def convert_docx(input_path, output_path, pages):
    """Convert selected pages of the PDF to a .docx file using pdf2docx."""
    # pdf2docx logs a lot (and some of its dependencies print to stdout), which would
    # break the single-JSON-line stdout contract. Silence its loggers and capture any
    # stray stdout writes during the conversion.
    for name in list(logging.root.manager.loggerDict.keys()) + [
        "pdf2docx",
        "fitz",
        "PyMuPDF",
    ]:
        logging.getLogger(name).setLevel(logging.CRITICAL)
    logging.disable(logging.CRITICAL)
    # Always restore logging — including if Converter() / import fails before convert().
    try:
        from pdf2docx import Converter

        cv = Converter(input_path)
        old_stdout = sys.stdout
        try:
            sys.stdout = io.StringIO()
            cv.convert(output_path, pages=pages)
        finally:
            sys.stdout = old_stdout
            cv.close()
    finally:
        logging.disable(logging.NOTSET)


def convert_xlsx(input_path, output_path, pages):
    """Convert selected pages of the PDF to a .xlsx workbook: one sheet per page.

    Each page's tables (via pdfplumber) become rows, blank row between multiple
    tables on the same page. Pages with no tables fall back to extracted text
    (one line per row). Fully empty pages get a single "(empty page)" cell.
    """
    import pdfplumber
    from openpyxl import Workbook

    wb = Workbook()
    # remove the default sheet created by openpyxl; we add our own per page
    wb.remove(wb.active)

    used_names = set()

    with pdfplumber.open(input_path) as pdf:
        for idx in pages:
            page = pdf.pages[idx]
            page_number_1based = idx + 1
            sheet_name = f"Page {page_number_1based}"
            if sheet_name in used_names:
                # shouldn't happen since indices are deduped, but guard anyway
                suffix = 2
                candidate = f"{sheet_name} ({suffix})"
                while candidate in used_names:
                    suffix += 1
                    candidate = f"{sheet_name} ({suffix})"
                sheet_name = candidate
            used_names.add(sheet_name)
            # Excel sheet names are capped at 31 chars
            sheet_name = sheet_name[:31]

            ws = wb.create_sheet(title=sheet_name)

            tables = page.extract_tables()
            if tables:
                for table_idx, table in enumerate(tables):
                    if table_idx > 0:
                        ws.append([])  # blank row between tables
                    for row in table:
                        ws.append(["" if cell is None else cell for cell in row])
            else:
                text = page.extract_text()
                if text:
                    for line in text.split("\n"):
                        ws.append([line])
                else:
                    ws["A1"] = "(empty page)"

    wb.save(output_path)


def run(mode, input_path, output_path, pages_arg):
    if not os.path.isfile(input_path):
        raise ConvertError(f"input file not found: {input_path}")

    if mode not in ("docx", "xlsx"):
        raise ConvertError(f"invalid mode '{mode}' (expected 'docx' or 'xlsx')")

    page_count = open_pdf_for_validation(input_path)
    pages = parse_pages_arg(pages_arg, page_count)

    out_dir = os.path.dirname(os.path.abspath(output_path))
    if out_dir and not os.path.isdir(out_dir):
        raise ConvertError(f"output directory does not exist: {out_dir}")

    if mode == "docx":
        convert_docx(input_path, output_path, pages)
    elif mode == "xlsx":
        convert_xlsx(input_path, output_path, pages)
    else:
        # unreachable: argparse restricts choices, but keep for safety
        raise ConvertError(f"unknown mode '{mode}'")

    return len(pages)


def main():
    parser = argparse.ArgumentParser(description="Convert a PDF to docx or xlsx.")
    # NOTE: mode is validated manually inside run() (not via argparse `choices`) so
    # an invalid mode produces our JSON error contract (exit 1) instead of argparse's
    # own usage-error output (exit 2).
    parser.add_argument("--mode", required=True)
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--pages", default=None)
    args = parser.parse_args()

    try:
        pages_converted = run(args.mode, args.input, args.output, args.pages)
    except ConvertError as exc:
        print(json.dumps({"ok": False, "error": str(exc)}), file=sys.stderr)
        sys.exit(1)
    except Exception as exc:
        # last resort: never let a raw traceback be the only error output.
        import traceback

        traceback.print_exc(file=sys.stderr)
        print(json.dumps({"ok": False, "error": f"unexpected error: {exc}"}), file=sys.stderr)
        sys.exit(1)

    print(json.dumps({"ok": True, "pages": pages_converted}))
    sys.exit(0)


if __name__ == "__main__":
    main()
