"""
Self-test for convert.py. Plain script, no pytest. Run with the project venv python:
    .venv\\Scripts\\python.exe test_convert.py

Generates a small 3-page test PDF, exercises convert.py via subprocess for both
success and error cases, and prints a PASS/FAIL summary line per case. Exits 0
only if every case passed.
"""

import json
import os
import subprocess
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
TESTOUT = os.path.join(HERE, ".testout")
CONVERT_PY = os.path.join(HERE, "convert.py")
PY = os.path.join(HERE, ".venv", "Scripts", "python.exe")

results = []  # list of (name, passed_bool, detail)


def record(name, ok, detail=""):
    results.append((name, ok, detail))
    status = "PASS" if ok else "FAIL"
    print(f"[{status}] {name}" + (f" - {detail}" if detail and not ok else ""))


def make_test_pdf(path):
    """Build a 3-page test PDF with pymupdf (fitz).

    page 1: paragraphs of text including CJK "中文測試"
    page 2: a drawn table (grid lines + cell text) for pdfplumber table detection
    page 3: mixed text + a small table
    """
    import fitz

    doc = fitz.open()

    # --- page 1: plain paragraphs, incl. CJK ---
    page1 = doc.new_page()
    # NOTE: this paragraph text is drawn with the base14 "helv" font, which has no CJK
    # glyphs. The CJK line below uses PyMuPDF's built-in "china-s" font instead, on a
    # separate insert_text call, so it round-trips through text extraction correctly.
    text1 = (
        "This is a test document for the PDF converter sidecar.\n"
        "It contains multiple paragraphs of plain text.\n\n"
        "Second paragraph with some more filler content to pad the page out\n"
        "so that extracted text is non-trivial in length.\n"
    )
    page1.insert_text((72, 72), text1, fontsize=11, fontname="helv")
    page1.insert_text((72, 220), "中文測試 中文測試 中文測試", fontsize=14, fontname="china-s")

    # --- page 2: drawn table (grid lines + text in cells) ---
    page2 = doc.new_page()
    rows, cols = 4, 3
    x0, y0 = 72, 100
    cell_w, cell_h = 100, 30
    # vertical lines
    for c in range(cols + 1):
        x = x0 + c * cell_w
        page2.draw_line((x, y0), (x, y0 + rows * cell_h))
    # horizontal lines
    for r in range(rows + 1):
        y = y0 + r * cell_h
        page2.draw_line((x0, y), (x0 + cols * cell_w, y))
    # cell text
    for r in range(rows):
        for c in range(cols):
            cx = x0 + c * cell_w + 8
            cy = y0 + r * cell_h + 20
            page2.insert_text((cx, cy), f"R{r}C{c}", fontsize=10, fontname="helv")

    # --- page 3: mixed text + small table ---
    page3 = doc.new_page()
    page3.insert_text((72, 72), "Mixed page: intro text followed by a small table.", fontsize=11)
    rows3, cols3 = 2, 2
    x0b, y0b = 72, 150
    cw, ch = 90, 25
    for c in range(cols3 + 1):
        x = x0b + c * cw
        page3.draw_line((x, y0b), (x, y0b + rows3 * ch))
    for r in range(rows3 + 1):
        y = y0b + r * ch
        page3.draw_line((x0b, y), (x0b + cols3 * cw, y))
    for r in range(rows3):
        for c in range(cols3):
            cx = x0b + c * cw + 6
            cy = y0b + r * ch + 17
            page3.insert_text((cx, cy), f"V{r}{c}", fontsize=10)

    doc.save(path)
    doc.close()


def run_convert(args):
    cmd = [PY, CONVERT_PY] + args
    proc = subprocess.run(cmd, capture_output=True, text=True, encoding="utf-8")
    return proc.returncode, proc.stdout, proc.stderr


def parse_json_line(text, stream_name, name):
    """Parse the LAST line of text as JSON. Returns dict or None on failure (records a FAIL)."""
    lines = [l for l in text.splitlines() if l.strip()]
    if not lines:
        record(name, False, f"no {stream_name} output at all")
        return None
    last = lines[-1]
    try:
        return json.loads(last)
    except json.JSONDecodeError as exc:
        record(name, False, f"last {stream_name} line is not valid JSON: {last!r} ({exc})")
        return None


def main():
    os.makedirs(TESTOUT, exist_ok=True)

    if not os.path.isfile(PY):
        print(f"FATAL: venv python not found at {PY}")
        sys.exit(1)

    pdf_path = os.path.join(TESTOUT, "test.pdf")
    make_test_pdf(pdf_path)
    record("generate test pdf", os.path.isfile(pdf_path) and os.path.getsize(pdf_path) > 0)

    # --- Case: docx all pages ---
    out_docx_all = os.path.join(TESTOUT, "all.docx")
    rc, out, err = run_convert(["--mode", "docx", "--input", pdf_path, "--output", out_docx_all])
    name = "docx all pages"
    if rc != 0:
        record(name, False, f"exit code {rc}, stderr: {err!r}")
    else:
        j = parse_json_line(out, "stdout", name)
        if j is not None:
            ok = (
                j.get("ok") is True
                and j.get("pages") == 3
                and os.path.isfile(out_docx_all)
                and os.path.getsize(out_docx_all) > 1024
            )
            detail = f"json={j}, size={os.path.getsize(out_docx_all) if os.path.isfile(out_docx_all) else 'missing'}"
            record(name, ok, detail)

    # --- Case: xlsx all pages ---
    out_xlsx_all = os.path.join(TESTOUT, "all.xlsx")
    rc, out, err = run_convert(["--mode", "xlsx", "--input", pdf_path, "--output", out_xlsx_all])
    name = "xlsx all pages"
    if rc != 0:
        record(name, False, f"exit code {rc}, stderr: {err!r}")
    else:
        j = parse_json_line(out, "stdout", name)
        if j is not None:
            ok = j.get("ok") is True and j.get("pages") == 3 and os.path.isfile(out_xlsx_all)
            if ok:
                from openpyxl import load_workbook

                wb = load_workbook(out_xlsx_all)
                expected_sheets = ["Page 1", "Page 2", "Page 3"]
                ok = wb.sheetnames == expected_sheets
                detail = f"sheetnames={wb.sheetnames}"
                if ok:
                    # sanity: page 2's sheet should have some non-empty rows (table or text fallback)
                    ws = wb["Page 2"]
                    non_empty_rows = sum(
                        1 for row in ws.iter_rows() if any(c.value not in (None, "") for c in row)
                    )
                    ok = non_empty_rows > 0
                    detail += f", page2_non_empty_rows={non_empty_rows}"
            else:
                detail = f"json={j}"
            record(name, ok, detail)

    # --- Case: docx pages 0,2 ---
    out_docx_subset = os.path.join(TESTOUT, "subset.docx")
    rc, out, err = run_convert(
        ["--mode", "docx", "--input", pdf_path, "--output", out_docx_subset, "--pages", "0,2"]
    )
    name = "docx pages 0,2"
    if rc != 0:
        record(name, False, f"exit code {rc}, stderr: {err!r}")
    else:
        j = parse_json_line(out, "stdout", name)
        if j is not None:
            ok = (
                j.get("ok") is True
                and j.get("pages") == 2
                and os.path.isfile(out_docx_subset)
                and os.path.getsize(out_docx_subset) > 1024
            )
            record(name, ok, f"json={j}")

    # --- Error case: bad mode ---
    rc, out, err = run_convert(
        ["--mode", "bogus", "--input", pdf_path, "--output", os.path.join(TESTOUT, "x1.docx")]
    )
    name = "error: bad mode"
    j = parse_json_line(err, "stderr", name)
    if j is not None:
        ok = rc == 1 and j.get("ok") is False and isinstance(j.get("error"), str) and j["error"]
        record(name, ok, f"rc={rc}, json={j}")

    # --- Error case: out-of-range page ---
    rc, out, err = run_convert(
        [
            "--mode",
            "docx",
            "--input",
            pdf_path,
            "--output",
            os.path.join(TESTOUT, "x2.docx"),
            "--pages",
            "0,12",
        ]
    )
    name = "error: out-of-range page"
    j = parse_json_line(err, "stderr", name)
    if j is not None:
        ok = (
            rc == 1
            and j.get("ok") is False
            and "out of range" in j.get("error", "")
            and "3 pages" in j.get("error", "")
        )
        record(name, ok, f"rc={rc}, json={j}")

    # --- Error case: duplicate page ---
    rc, out, err = run_convert(
        [
            "--mode",
            "docx",
            "--input",
            pdf_path,
            "--output",
            os.path.join(TESTOUT, "x3.docx"),
            "--pages",
            "1,1",
        ]
    )
    name = "error: duplicate page"
    j = parse_json_line(err, "stderr", name)
    if j is not None:
        ok = rc == 1 and j.get("ok") is False and "duplicate" in j.get("error", "").lower()
        record(name, ok, f"rc={rc}, json={j}")

    # --- Error case: missing input ---
    missing_path = os.path.join(TESTOUT, "does_not_exist.pdf")
    rc, out, err = run_convert(
        ["--mode", "docx", "--input", missing_path, "--output", os.path.join(TESTOUT, "x4.docx")]
    )
    name = "error: missing input"
    j = parse_json_line(err, "stderr", name)
    if j is not None:
        ok = rc == 1 and j.get("ok") is False and "not found" in j.get("error", "").lower()
        record(name, ok, f"rc={rc}, json={j}")

    # --- summary ---
    print()
    print("=== SUMMARY ===")
    all_ok = True
    for name, ok, detail in results:
        status = "PASS" if ok else "FAIL"
        print(f"{status}: {name}")
        if not ok:
            all_ok = False

    if all_ok:
        print("\nALL TESTS PASSED")
        sys.exit(0)
    else:
        print("\nSOME TESTS FAILED")
        sys.exit(1)


if __name__ == "__main__":
    main()
